from openpyxl import load_workbook
from player import Player

import twitter

# Column number constants zero indexed
NAME = 0
TEAM = 1
POSITION = 2
CITY = 3
STATE = 4
ACCESS_KEY = "1242300392-mWfwumCtmk1S9wwE85pwpF1oh8Gudqi8JiO0SWX"
ACCESS_SECRET = "T0bzmNJjOIeu1OGdGjel9KrsdC6npTdRd5NkZtJWCih3M"
CONSUMER_KEY = "Gu3fyUvzNFSQYtj2SV8bpDg93"
CONSUMER_SECRET = "EoQ9HFLOZRAqCuQmSYX3bEFZPndaw3TlOFiRBqsUEC5l61lDGE"

def get_twitter(player):
    api = twitter.Api(consumer_key=CONSUMER_KEY,
                  consumer_secret=CONSUMER_SECRET,
                  access_token_key=ACCESS_KEY,
                  access_token_secret=ACCESS_SECRET)
    results = api.GetUsersSearch(term=player.name)
    print(len(results))
    if len(results) > 0:
        print("https://twitter.com/intent/user?user_id=" + str(results[0].id))
        
    
    # for result in results:
    #     print(result["screen_name"])

def main():
    # Get the first workseet in the excel file
    wb = load_workbook('unlRoster.xlsx', read_only=True, data_only=True)
    ws = wb._sheets[0]

    # Get a players info from each row
    # The + 1 accounts for the headers
    # TODO: Find a better way to get max collumn
    players = []
    for row in ws['A{}:Z{}'.format(ws.min_row + 1, ws.max_row)]:
        player = Player(row[NAME].value, row[TEAM].value, row[POSITION].value, row[CITY].value, row[STATE].value)
        players.append(player)

    for player in players:
        get_twitter(player)

if __name__ == "__main__":
    main()
