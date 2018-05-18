from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .forms import RosterForm, RosterNameForm, RosterMinForm, RosterFromUrl, TwitterSeachTermForm
from .models import Roster, Player, TwitterResult, Friend
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup
from urllib.request import urlopen
from django.conf import settings

import tweepy

AK = '1589257088-VhcKzZDFDALPgNMTgY3aUX5N5hFC0S7HAPBot0C'
AS = 'XSlLMTKuRl0shTMEmnLHuUu397OoUyOX8zGURpkBLyDKw'
CK = '9OClm0WPv0AiDHoQ7ih7JI011'
CS = 'B37NdXd4zQR4PGdI3Q5OCdZj2DB6w6ctqMKgJdpgZToDkL1esG'


# The landing page containing a vew of all of your rosters and new roster form
def index(request):
    # This same format is used throughout, method == 'POST' tells you if
    # if someone is submitting a form
    if request.method == 'POST':
        upload_form = RosterForm(request.POST, request.FILES)
        if upload_form.is_valid():
            # Add the name of the file to the roster name before you save it
            roster = upload_form.save(commit=False)
            roster.name = str(roster.document).replace('.xlsx', '')
            roster.save()

            # Get the first worksheet in the excel file
            wb = load_workbook(roster.document, read_only=True, data_only=True)
            ws = wb._sheets[0]

            name_int = col_to_int(roster.name_col)

            # Add the players to the roster
            for row in ws['A{}:Z{}'.format(ws.min_row + 1, ws.max_row)]:
                if row[name_int].value is not None:
                    player = Player(name=row[name_int].value, roster=roster)
                    player.save()
            return redirect('roster', pk=roster.pk)
    else:
        upload_form = RosterForm()

    # Get all rosters
    rosters = Roster.objects.all()

    return render(request, 'frontEnd/index.html', {'form': upload_form, 'rosters': rosters})


# The detail page for the rosters
def roster_details(request, pk):
    # These two variables tell you if they are editing the name or min
    # values and if they are it makes the form appear
    is_editing_name = request.GET.get('is_editing_name', 'False')
    is_editing_min = request.GET.get('is_editing_min', 'False')

    roster = get_object_or_404(Roster, pk=pk)
    players = roster.player_set.all()

    # If the name form is being edited make it appear and handle if
    # there is a submission
    if is_editing_name == 'True':
        if request.method == 'POST':
            name_form = RosterNameForm(request.POST, instance=roster)
            if name_form.is_valid():
                roster = name_form.save(commit=False)
                roster.name = roster.name
                roster.save()
                return redirect('roster', pk=roster.pk)
        else:
            name_form = RosterNameForm(instance=roster)
    else:
        name_form = None

    # If the name form is being edited make it appear and handle if
    # there is a submission
    if is_editing_min == 'True':
        if request.method == 'POST':
            min_form = RosterMinForm(request.POST, instance=roster)
            if min_form.is_valid():
                roster = min_form.save(commit=False)
                roster.min_followers = roster.min_followers
                roster.save()
                return redirect('roster', pk=roster.pk)
        else:
            min_form = RosterMinForm(instance=roster)
    else:
        min_form = None

    # Delete all of the search results from the roster otherwise they would
    # build up like crazy
    for player in players:
        for result in player.twitterresult_set.all():
            result.delete()

    return render(request, 'frontEnd/roster.html', {'players': players,
                                                    'roster': roster,
                                                    'name_form': name_form,
                                                    'is_editing_name': is_editing_name,
                                                    'min_form': min_form,
                                                    'is_editing_min': is_editing_min})


# View for deleting a roster, returns to index when finished
def remove_roster(request, pk):
    roster = get_object_or_404(Roster, pk=pk)
    roster.delete()
    return redirect('index')


# View for downloading roster
def download_roster(request, pk):
    roster = get_object_or_404(Roster, pk=pk)

    # Loads the existing workbook and sheet
    wb = load_workbook(roster.document)
    ws = wb._sheets[0]

    # Set header cell
    cell = roster.twitter_col + '1'
    ws[cell] = 'Twitter'

    # Loop through the players and adds their name and twitter accounts
    i = 2
    for player in roster.player_set.all():
        ws[roster.twitter_col + str(i)] = player.twitter
        i += 1

    # Adds the excel file to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + roster.name + '.xlsx'
    wb.save(response)

    return response


# View for finding twitter account of a player
def find_twitter(request, pk):
    player = get_object_or_404(Player, pk=pk)

    # The search term form allows you to change the players name
    # without saving it, effectively changing the search term
    if request.method == 'POST':
        form = TwitterSeachTermForm(request.POST, instance=player)
        if form.is_valid():
            player = form.save(commit=False)
    else:
        form = TwitterSeachTermForm(instance=player)

    # The get twitter function is responsible for adding search results to a player
    get_twitter(player)

    return render(request, 'frontEnd/findTwitter.html', {'player': player, 'form': form})


# View that handles adding a twitter account to a player
def select_twitter_result(request, pk):
    result = get_object_or_404(TwitterResult, pk=pk)
    player = result.player
    player.twitter = result.screen_name
    player.twitter_id = result.user_id
    player.save()

    # Function that adds all of a players friends to the roster so they
    # can be used to enhance search results
    add_followers(player)

    return redirect('roster', pk=player.roster.pk)


# Sets a players twitter to "NONE" this is different than deleting it because it
# is saying that this players twitter does not exist, not that it hasn't been
# found yet
def no_twitter(request, pk):
    player = get_object_or_404(Player, pk=pk)
    player.twitter = 'NONE'
    player.twitter_id = ''
    player.save()
    return redirect('roster', pk=player.roster.pk)


# Removes twitter information from a player
def remove_twitter(request, pk):
    player = get_object_or_404(Player, pk=pk)
    player.twitter = ''
    player.twitter_id = ''
    player.save()
    return redirect('roster', pk=player.roster.pk)


# View that takes a ESPN url and scrapes the roster
def roster_from_url(request):
    if request.method == 'POST':
        form = RosterFromUrl(request.POST, request.FILES)
        if form.is_valid():
            # Create excel file and make the roster sheet
            wb = Workbook()
            ws = wb.active
            ws.title = 'Roster'

            # Use beautiful soup to get html
            html = urlopen(form.cleaned_data['url'])
            soup = BeautifulSoup(html, 'html.parser')

            # Find the table and then get the title and search the rows
            title = soup.find('h1', {'class': 'h2'}).find(text=True)
            table = soup.find('table', {'class': 'tablehead'})
            rows = table.find_all('tr')

            # Make a roster so we can add players to it
            roster = Roster(name_col='B', twitter_col='H')
            roster.save()

            # Get the data out of the table and into a matrix
            rows_data = []
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 7:
                    row = [cells[0].find(text=True),
                           cells[1].find(text=True),
                           cells[2].find(text=True),
                           cells[3].find(text=True),
                           cells[4].find(text=True),
                           cells[5].find(text=True),
                           cells[6].find(text=True)]
                    rows_data.append(row)
                    if cells[1].find(text=True) != 'NAME':
                        player = Player(roster=roster, name=cells[1].find(text=True))
                        player.save()

            # Put that data into the spreadsheet
            i = 0
            j = 0
            for sheet_row in ws.iter_rows(min_row=1, max_col=7, max_row=len(rows_data)):
                for cell in sheet_row:
                    cell.value = rows_data[i][j]
                    j += 1
                j = 0
                i += 1

            # Save the excel file to the proper folder
            wb.save(settings.MEDIA_ROOT + '/documents/' + title + '.xlsx')

            # Add the title and file to the roster
            roster.name = title
            roster.document.name = 'documents/' + title + '.xlsx'
            roster.save()

            return redirect('roster', pk=roster.pk)
    else:
        form = RosterFromUrl()

    return render(request, 'frontEnd/rosterFromUrl.html', {'form': form})


# Function that takes a player and adds twitter results
def get_twitter(player):
    # Clear old results if there are any
    for result in player.twitterresult_set.all():
        result.delete()

    # Authenticate with tweepy
    auth = tweepy.OAuthHandler(CK, CS)
    auth.set_access_token(AK, AS)
    api = tweepy.API(auth)

    # Get the search results of the player name
    search_results = api.search_users(q=player.name)

    # This stops an error from occurring when there are no results
    if len(search_results) == 0:
        return

    # Process each search result
    for search_result in search_results:
        # Ignore anything bellow the minimum follower count to filter out
        # worthless accounts
        if search_result.followers_count < player.roster.min_followers:
            continue

        # Try catch because sometimes the banner url doesn't exist, when
        # it is set to non a default is provided in models.py
        try:
            profile_banner_url = search_result.profile_banner_url
        except:
            profile_banner_url = None

        # The .replace("_normal","") changes the url for the profile image so we get the higher quality version
        result = TwitterResult(player=player,
                               name=search_result.name,
                               screen_name=search_result.screen_name,
                               user_id=search_result.id,
                               description=search_result.description,
                               profile_banner_url=profile_banner_url,
                               profile_image_url=search_result.profile_image_url.replace('_normal', ''))

        # Goes through every follower of players on the roster and if the result has the same id as one
        # of them it marks them as a friend
        if any(follower.user_id == result.user_id for follower in result.player.roster.friend_set.all()):
            result.friend = True

        result.save()


# Adds the followers of a player so they can be used to mark players friends in the results


def add_followers(player):
    # Authenticate with tweepy
    auth = tweepy.OAuthHandler(CK, CS)
    auth.set_access_token(AK, AS)
    api = tweepy.API(auth)

    # One time I got an error so I put this in a try catch
    try:
        # Creates a friend for each person the player follows
        for follower_id in tweepy.Cursor(api.friends_ids, screen_name=player.twitter).items():
            follower = Friend(player=player, roster=player.roster, user_id=follower_id)
            follower.save()
    except tweepy.TweepError:
        print("Can't do it")


# A function to convert a column letter into a zero index int
#
# Ex) A -> 0, B -> 1
def col_to_int(col):
    col = col.upper()
    return ord(col) - 65
